import openpyxl
from openpyxl.styles import Font,colors

inv_file = openpyxl.load_workbook("sheet1.xlsx")
emp_list = inv_file["Sheet1"]

bold_and_red_font = Font(bold=True, color='FA6075')
employees_per_county = {}
total_country_value = {}
employees_under_80k = {}
total_emp_wealth_label = emp_list.cell(1, 11)
#print(emp_list.max_row)

for employee_row in range(2, emp_list.max_row + 1):
    country = emp_list.cell(employee_row, 5).value
    annual_salary = emp_list.cell(employee_row, 6).value
    years_of_service = emp_list.cell(employee_row, 7).value
    employee_name = emp_list.cell(employee_row, 2).value
    total_emp_wealth = emp_list.cell(employee_row, 11)
    #print(county)
    # calculation for number of employees per country
    if country in employees_per_county:
        current_num_employees = employees_per_county.get(country)
        employees_per_county[country] = current_num_employees + 1
    else:
        employees_per_county[country] = 1

    # calculating total wealth for each county employees
    if country in total_country_value:
        current_total_value = total_country_value.get(country)
        total_country_value[country] = current_total_value + annual_salary * years_of_service
    else:
        total_country_value[country] = annual_salary * years_of_service

    # logic to list employees whose salary is less than 100K
    if annual_salary < 80000:
        employees_under_80k[employee_name] = annual_salary

    # adding total wealth for each employee
    if employee_row == 2:
        total_emp_wealth_label.value = "Emp Wealth"
        total_emp_wealth_label.font = bold_and_red_font
        total_emp_wealth.value = annual_salary * years_of_service
    else:
        total_emp_wealth.value = annual_salary * years_of_service

print(employees_per_county)
print(total_country_value)
print(employees_under_80k)
inv_file.save("sheet1_with_total_emp_wealth.xlsx")
#testing Github sync




